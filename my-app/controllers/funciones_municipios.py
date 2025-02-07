# Python con PostgreSQL
# Para subir archivo tipo foto al servidor
from werkzeug.utils import secure_filename
import uuid  # Modulo de python para crear un string

from conexion.conexionBD import connectionBD  # Conexión a BD


import datetime
import re
import os

from os import remove  # Modulo  para remover archivo
from os import path  # Modulo para obtener la ruta o directorio

import openpyxl  # Para generar el excel
# biblioteca o modulo send_file para forzar la descarga
from flask import send_file

# Importar DictCursor para obtener resultados como diccionarios
from psycopg2.extras import DictCursor


# ********************************************************** sección de municipios **********************************************************

def procesar_imagen_perfil(foto):
    try:
        # Nombre original del archivo
        filename = secure_filename(foto.filename)
        extension = os.path.splitext(filename)[1]

        # Creando un string de 50 caracteres
        nuevoNameFile = (uuid.uuid4().hex + uuid.uuid4().hex)[:100]
        nombreFile = nuevoNameFile + extension

        # Construir la ruta completa de subida del archivo
        basepath = os.path.abspath(os.path.dirname(__file__))
        upload_dir = os.path.join(basepath, f'../static/fotos_municipios/')

        # Validar si existe la ruta y crearla si no existe
        if not os.path.exists(upload_dir):
            os.makedirs(upload_dir)
            # Dando permiso a la carpeta
            os.chmod(upload_dir, 0o755)

        # Construir la ruta completa de subida del archivo
        upload_path = os.path.join(upload_dir, nombreFile)
        foto.save(upload_path)

        return nombreFile

    except Exception as e:
        print("Error al procesar archivo:", e)
        return []


# Lista de municipios - PostgreSQL 
def sql_lista_municipiosBD():
    try:
        with connectionBD() as conexion_PostgreSQLdb:
            with conexion_PostgreSQLdb.cursor(cursor_factory=DictCursor) as cursor:
                querySQL = """
                    SELECT 
                        um.id,
                        um.region, 
                        um.departamento,
                        um.municipio,
                        um.sup_parcelas,
                        um.num_parcelas,
                        um.sup_muestra,                      
                        CASE
                            WHEN um.seleccionado = 1 THEN 'Seleccionado'
                            ELSE 'No Seleccionado'
                        END AS seleccion_municipio
                    FROM unodc_municipios AS um
                    ORDER BY um.id DESC
                """
                cursor.execute(querySQL)
                municipiosBD = cursor.fetchall()
        return municipiosBD
    except Exception as e:
        print(f"Error en la función sql_lista_municipiosBD: {e}")
        return None


# Detalles del municipio - PostgreSQL 
def sql_detalles_municipiosBD(idmunicipio):
    try:
        with connectionBD() as conexion_PostgreSQLdb:
            with conexion_PostgreSQLdb.cursor(cursor_factory=DictCursor) as cursor:
                querySQL = """
                    SELECT 
                        um.id,
                        um.region, 
                        um.departamento,
                        um.municipio,
                        um.sup_parcelas,
                        um.num_parcelas,
                        um.sup_muestra,                      
                        CASE
                            WHEN um.seleccionado = 1 THEN 'Seleccionado'
                            ELSE 'No Seleccionado'
                        END AS seleccion_municipio
                    FROM unodc_municipios AS um
                    WHERE id = %s
                    ORDER BY um.id DESC
                """
                cursor.execute(querySQL, (idmunicipio,))
                municipiosBD = cursor.fetchone()
        return municipiosBD
    except Exception as e:
        print(f"Error en la función sql_detalles_municipiosBD: {e}")
        return None


# Funcion municipios Informe (Reporte) - PostgreSQL
def municipiosReporte():
    try:
        with connectionBD() as conexion_PostgreSQLdb:
            with conexion_PostgreSQLdb.cursor(cursor_factory=DictCursor) as cursor:
                querySQL = """
                    SELECT 
                        um.id,
                        um.municipio, 
                        um.region,
                        um.departamento,
                        um.sup_parcelas,
                        um.num_parcelas,
                        um.sup_muestra
                    FROM unodc_municipios AS um
                    WHERE um.seleccionado = 1
                    ORDER BY um.id DESC
                """
                cursor.execute(querySQL)
                municipiosBD = cursor.fetchall()
        return municipiosBD
    except Exception as e:
        print(f"Error en la función municipiosReporte: {e}")
        return None

# Funcion municipios Generar Reporte Excel - PostgreSQL
def generarReporteExcelxmun():
    dataMunicipios = municipiosReporte()
    wb = openpyxl.Workbook()
    hoja = wb.active

    # Agregar la fila de encabezado con los títulos
    cabeceraExcel = ("Municipio", "Region", "Departamento",
                     "Sup Parcelas", "Num Parcelas", "Sup Muestra")

    hoja.append(cabeceraExcel)

    # Formato para números en moneda colombiana y sin decimales
    formato_moneda_colombiana = '#,##0'

    # Agregar los registros a la hoja
    for registro in dataMunicipios:
        municipio = registro['municipio']
        region = registro['region']
        departamento = registro['departamento']
        sup_parcelas = registro['sup_parcelas']
        num_parcelas = registro['num_parcelas']
        sup_muestra = registro['sup_muestra']
        
        # Agregar los valores a la hoja
        hoja.append((municipio, region, departamento, sup_parcelas, num_parcelas,sup_muestra))

        # Itera a través de las filas y aplica el formato a la columna G
        for fila_num in range(2, hoja.max_row + 1):
            columna = 7  # Columna G
            celda = hoja.cell(row=fila_num, column=columna)
            celda.number_format = formato_moneda_colombiana

    fecha_actual = datetime.datetime.now()
    archivoExcel = f"Reporte_municipios_{fecha_actual.strftime('%Y_%m_%d')}.xlsx"
    carpeta_descarga = "../static/downloads-excel"
    ruta_descarga = os.path.join(os.path.dirname(
        os.path.abspath(__file__)), carpeta_descarga)

    if not os.path.exists(ruta_descarga):
        os.makedirs(ruta_descarga)
        # Dando permisos a la carpeta
        os.chmod(ruta_descarga, 0o755)

    ruta_archivo = os.path.join(ruta_descarga, archivoExcel)
    wb.save(ruta_archivo)

    # Enviar el archivo como respuesta HTTP
    return send_file(ruta_archivo, as_attachment=True)

# Buscar municipios Unico - PostgreSQL 
def buscarMunicipioBD(search):
    try:
        with connectionBD() as conexion_PostgreSQLdb:
            with conexion_PostgreSQLdb.cursor(cursor_factory=DictCursor) as cursor:
                querySQL = """
                    SELECT 
                        um.id,
                        um.region, 
                        um.departamento,
                        um.municipio,
                        um.sup_parcelas,
                        um.num_parcelas,
                        um.sup_muestra,                      
                        CASE
                            WHEN um.seleccionado = 1 THEN 'Seleccionado'
                            ELSE 'No Seleccionado'
                        END AS seleccion_municipio
                    FROM unodc_municipios AS um
                    
                    WHERE um.municipio ILIKE %s 
                    ORDER BY um.id DESC
                """
                search_pattern = f"%{search}%"  # Agregar "%" alrededor del término de búsqueda
                cursor.execute(querySQL, (search_pattern,))
                resultado_busqueda = cursor.fetchall()
                return resultado_busqueda

    except Exception as e:
        print(f"Ocurrió un error en def buscarMunicipioBD: {e}")
        return []

# Buscar municipio Unico (cuando viene del grid) - PostgreSQL XXXXXXXXXXXXXX
def buscarMunicipioUnico(id):
    try:
        with connectionBD() as conexion_PostgreSQLdb:
            with conexion_PostgreSQLdb.cursor(cursor_factory=DictCursor) as cursor:
                querySQL = """
                    SELECT 
                        um.id,
                        um.region, 
                        um.departamento,
                        um.municipio,
                        um.sup_parcelas,
                        um.num_parcelas,
                        um.sup_muestra,                      
                        CASE
                            WHEN um.seleccionado = 1 THEN 'Seleccionado'
                            ELSE 'No Seleccionado'
                        END AS seleccion_municipio
                    FROM unodc_municipios AS um
                    WHERE um.id = %s LIMIT 1
                """
                cursor.execute(querySQL, (id,))
                municipio = cursor.fetchone()
                return municipio

    except Exception as e:
        print(f"Ocurrió un error en def buscarMunicipioUnico: {e}")
        return []


def procesar_actualizacion_form(data):
    try:
        with connectionBD() as conexion_PostgreSQLdb:
            with conexion_PostgreSQLdb.cursor(cursor_factory=DictCursor) as cursor:
                municipio = data.form['municipio']
                region = data.form['region']
                departamento = data.form['departamento']
                sup_parcelas = data.form['sup_parcelas']
                num_parcelas = data.form['num_parcelas']
                sup_muestra = data.form['sup_muestra']
                id = data.form['id']
                
                querySQL = """
                        UPDATE unodc_municipios
                        SET 
                            municipio = %s,
                            region = %s,
                            departamento = %s,
                            sup_parcelas = %s,
                            num_parcelas = %s,
                            sup_muestra = %s
                        WHERE id = %s
                """
                values = (municipio, region, departamento, sup_parcelas, num_parcelas, sup_muestra, id)

                cursor.execute(querySQL, values)
                conexion_PostgreSQLdb.commit()

        return cursor.rowcount or []
    except Exception as e:
        print(f"Ocurrió un error en procesar_actualizacion_form: {e}")
        return None

def resetearTablas():
    try:
        with connectionBD() as conexion_PostgreSQLdb:
            with conexion_PostgreSQLdb.cursor() as cursor:
                # Poner en cero las tablas
                queries = [
                    "UPDATE unodc_municipios SET sup_parcelas = 0, num_parcelas = 0, sorteado = 0, seleccionado = 0, sup_muestra = 0, num_muestra = 0 WHERE TRUE",
                    "UPDATE unodc_parcelas SET sorteado = 0, seleccionado = 0, mun_seleccionado = 0, sup_muestra = 0, num_muestra = 0 WHERE TRUE",
                    "UPDATE unodc_municipios SET sup_parcelas = (SELECT SUM(sup_parcelas) FROM unodc_parcelas WHERE unodc_municipios.cod_mun = unodc_parcelas.cod_mun)",
                    "UPDATE unodc_municipios SET num_parcelas = (SELECT SUM(num_parcelas) FROM unodc_parcelas WHERE unodc_municipios.cod_mun = unodc_parcelas.cod_mun)",
                    "UPDATE unodc_municipios SET sorteado = (SELECT COUNT(id) FROM unodc_parcelas WHERE unodc_municipios.cod_mun = unodc_parcelas.cod_mun)",
                    "UPDATE unodc_municipios SET seleccionado = 1 WHERE sorteado >= 1",
                    "UPDATE unodc_municipios SET sup_muestra = sup_parcelas * 0.10, num_muestra = num_parcelas * 0.10 WHERE seleccionado = 1",
                    """
                    UPDATE unodc_parcelas t1
                    SET mun_seleccionado = 1
                    FROM unodc_municipios t2
                    WHERE t1.cod_mun = t2.cod_mun AND t2.seleccionado = 1
                    """
                ]
                for query in queries:
                    cursor.execute(query)
                conexion_PostgreSQLdb.commit()

    except Exception as e:
        print(f"Ocurrió un error en def resetearTablas: {e}")

