# Python con PostgreSQL
# Para subir archivo tipo foto al servidor
from werkzeug.utils import secure_filename
import uuid  # Modulo de python para crear un string

from conexion.conexionBD import connectionBD  # Conexión a BD


import datetime
import re
import os

from os import remove  # Modulo  para remover archivo
from os import path  # Modulo para obtener la ruta o directorio

import openpyxl  # Para generar el excel
# biblioteca o modulo send_file para forzar la descarga
from flask import send_file

# Importar DictCursor para obtener resultados como diccionarios
from psycopg2.extras import DictCursor


# ********************************************************** sección de puntos **********************************************************

def procesar_form_empleado(dataForm, foto_perfil):
    # Formateando Salario
    sup_muestra_sin_puntos = re.sub('[^0-9]+', '', dataForm['sup_muestra'])
    # convertir salario a INT
    sup_muestra_entero = int(sup_muestra_sin_puntos)

    result_foto_perfil = procesar_imagen_perfil(foto_perfil)
    try:
        with connectionBD() as conexion_PostgreSQLdb:
            with conexion_PostgreSQLdb.cursor(cursor_factory=DictCursor) as cursor:

                sql = """
                    INSERT INTO parcelas_registro 
                        (region, departamento, municipio, cod_hex, sup_parcelas, num_parcelas, densidad, sup_muestra, num_muestra, seleccionado, foto_muestra) 
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """

                # Creando una tupla con los valores del INSERT
                valores = (
                    dataForm['region'], dataForm['departamento'], dataForm['municipio'], dataForm['cod_hex'],
                    dataForm['sup_parcelas'], dataForm['num_parcelas'], dataForm['densidad'], dataForm['sup_muestra'],
                    dataForm['num_muestra'], dataForm['seleccionado'],
                    result_foto_perfil, sup_muestra_entero
                )
                cursor.execute(sql, valores)

                conexion_PostgreSQLdb.commit()
                resultado_insert = cursor.rowcount
                return resultado_insert

    except Exception as e:
        return f'Se produjo un error en procesar_form_empleado: {str(e)}'


def procesar_imagen_perfil(foto):
    try:
        # Nombre original del archivo
        filename = secure_filename(foto.filename)
        extension = os.path.splitext(filename)[1]

        # Creando un string de 50 caracteres
        nuevoNameFile = (uuid.uuid4().hex + uuid.uuid4().hex)[:100]
        nombreFile = nuevoNameFile + extension

        # Construir la ruta completa de subida del archivo
        basepath = os.path.abspath(os.path.dirname(__file__))
        upload_dir = os.path.join(basepath, f'../static/fotos_puntos/')

        # Validar si existe la ruta y crearla si no existe
        if not os.path.exists(upload_dir):
            os.makedirs(upload_dir)
            # Dando permiso a la carpeta
            os.chmod(upload_dir, 0o755)

        # Construir la ruta completa de subida del archivo
        upload_path = os.path.join(upload_dir, nombreFile)
        foto.save(upload_path)

        return nombreFile

    except Exception as e:
        print("Error al procesar archivo:", e)
        return []
    
# Lista de puntos - PostgreSQL 
def sql_lista_puntosBD():
    try:
        with connectionBD() as conexion_PostgreSQLdb:
            with conexion_PostgreSQLdb.cursor(cursor_factory=DictCursor) as cursor:
                querySQL = """
                    SELECT 
                        pr.id,
                        pr.region, 
                        pr.departamento,
                        pr.municipio,
                        pr.cod_hex,
                        pr.sup_parcelas,
                        pr.num_parcelas,
                        pr.densidad,
                        pr.sup_muestra,                      
                        pr.num_muestra,
                        pr.foto,
                        CASE
                            WHEN pr.seleccionado = 1 THEN 'Seleccionado'
                            ELSE 'No Seleccionado'
                        END AS seleccion_punto
                    FROM parcelas_registro AS pr 
                    WHERE pr.seleccionado = 1
                    ORDER BY pr.id DESC
                """
                cursor.execute(querySQL)
                puntosBD = cursor.fetchall()
        return puntosBD
    except Exception as e:
        print(f"Error en la función sql_lista_puntosBD:{e}")
        return None


# Detalles del punto - PostgreSQL 
def sql_detalles_puntosBD(idpunto):
    try:
        with connectionBD() as conexion_PostgreSQLdb:
            with conexion_PostgreSQLdb.cursor(cursor_factory=DictCursor) as cursor:
                querySQL = """
                    SELECT 
                        pr.id,
                        pr.region, 
                        pr.departamento,
                        pr.municipio,
                        pr.cod_hex,
                        pr.sup_parcelas,
                        pr.num_parcelas,
                        pr.densidad,
                        pr.sup_muestra,                      
                        pr.num_muestra,
                        CASE
                            WHEN pr.seleccionado = 1 THEN 'Seleccionado'
                            ELSE 'No Seleccionado'
                        END AS seleccion_punto
                    FROM parcelas_registro AS pr 
                    WHERE pr.id = %s
                    ORDER BY pr.id DESC
                """
                cursor.execute(querySQL, (idpunto,))
                puntosBD = cursor.fetchone()
        return puntosBD
    except Exception as e:
        print(f"Error en la función sql_detalles_puntosBD:{e}")
        return None


# Funcion puntos Informe (Reporte) - PostgreSQL
def puntosReporte():
    try:
        with connectionBD() as conexion_PostgreSQLdb:
            with conexion_PostgreSQLdb.cursor(cursor_factory=DictCursor) as cursor:
                querySQL = """
                    SELECT 
                        pr.id,
                        pr.region, 
                        pr.departamento,
                        pr.municipio,
                        pr.cod_hex,
                        pr.sup_parcelas,
                        pr.num_parcelas,
                        pr.densidad,
                        pr.sup_muestra,                      
                        pr.num_muestra,
                        CASE
                            WHEN pr.seleccionado = 1 THEN 'Seleccionado'
                            ELSE 'No Seleccionado'
                        END AS seleccion_punto
                    FROM parcelas_registro AS pr 
                    WHERE pr.seleccionado = 1
                    ORDER BY pr.id DESC
                """
                cursor.execute(querySQL)
                puntosBD = cursor.fetchall()
        return puntosBD
    except Exception as e:
        print(f"Error en la función puntosReporte:{e}")
        return None

# Funcion puntos Generar Reporte Excel - PostgreSQL
def generarReporteExcelxhex():
    dataPuntos = puntosReporte()
    wb = openpyxl.Workbook()
    hoja = wb.active

    # Agregar la fila de encabezado con los títulos
    cabeceraExcel = ( "Region", "Departamento", "Municipio", "Cod Hex", "Sup Parcelas", "Num Parcelas", "Densidad", "Sup Muestra", "Num Muestra", "Seleccionado")
                     

    hoja.append(cabeceraExcel)

    # Formato para números en moneda colombiana y sin decimales
    formato_moneda_colombiana = '#,##0'

    # Agregar los registros a la hoja
    for registro in dataPuntos:
        region = registro['region']
        departamento = registro['departamento']
        municipio = registro['municipio']
        sup_parcelas = registro['sup_parcelas']
        num_parcelas = registro['num_parcelas']
        densidad = registro['densidad']
        sup_muestra = registro['sup_muestra']
        num_muestra = registro['num_muestra']
        seleccion_punto = registro['seleccion_punto']
        
        # Agregar los valores a la hoja
        hoja.append((region, departamento, municipio, sup_parcelas, num_parcelas, densidad, sup_muestra, num_muestra, seleccion_punto))

        # Itera a través de las filas y aplica el formato a la columna G
        for fila_num in range(2, hoja.max_row + 1):
            columna = 7  # Columna G
            celda = hoja.cell(row=fila_num, column=columna)
            celda.number_format = formato_moneda_colombiana

    fecha_actual = datetime.datetime.now()
    archivoExcel = f"Reporte_puntos_{fecha_actual.strftime('%Y_%m_%d')}.xlsx"
    carpeta_descarga = "../static/downloads-excel"
    ruta_descarga = os.path.join(os.path.dirname(
        os.path.abspath(__file__)), carpeta_descarga)

    if not os.path.exists(ruta_descarga):
        os.makedirs(ruta_descarga)
        # Dando permisos a la carpeta
        os.chmod(ruta_descarga, 0o755)

    ruta_archivo = os.path.join(ruta_descarga, archivoExcel)
    wb.save(ruta_archivo)

    # Enviar el archivo como respuesta HTTP
    return send_file(ruta_archivo, as_attachment=True)

# Buscar puntos Unico - PostgreSQL 
def buscarPuntoBD(search):
    try:
        with connectionBD() as conexion_PostgreSQLdb:
            with conexion_PostgreSQLdb.cursor(cursor_factory=DictCursor) as cursor:
                querySQL = """
                    SELECT 
                        pr.id,
                        pr.region, 
                        pr.departamento,
                        pr.municipio,
                        pr.cod_hex,
                        pr.sup_parcelas,
                        pr.num_parcelas,
                        pr.densidad,
                        pr.sup_muestra,                      
                        pr.num_muestra,
                        CASE
                            WHEN pr.seleccionado = 1 THEN 'Seleccionado'
                            ELSE 'No Seleccionado'
                        END AS seleccion_punto
                    FROM parcelas_registro AS pr 
                    WHERE pr.cod_hex ILIKE %s 
                    ORDER BY pr.id DESC
                """
                search_pattern = f"%{search}%"  # Agregar "%" alrededor del término de búsqueda
                cursor.execute(querySQL, (search_pattern,))
                resultado_busqueda = cursor.fetchall()
                return resultado_busqueda

    except Exception as e:
        print(f"Ocurrió un error en def buscarPuntoBD:{e}")
        return []

# Buscar punto Unico (cuando viene del grid) - PostgreSQL XXXXXXXXXXXXXX
def buscarPuntoUnico(id):
    try:
        with connectionBD() as conexion_PostgreSQLdb:
            with conexion_PostgreSQLdb.cursor(cursor_factory=DictCursor) as cursor:
                querySQL = """
                    SELECT 
                        pr.id,
                        pr.region, 
                        pr.departamento,
                        pr.municipio,
                        pr.cod_hex,
                        pr.sup_parcelas,
                        pr.num_parcelas,
                        pr.densidad,
                        pr.sup_muestra,                      
                        pr.num_muestra,
                        CASE
                            WHEN pr.seleccionado = 1 THEN 'Seleccionado'
                            ELSE 'No Seleccionado'
                        END AS seleccion_punto
                    FROM parcelas_registro AS pr 
                    WHERE pr.id = %s LIMIT 1
                """
                cursor.execute(querySQL, (id,))
                punto = cursor.fetchone()
                return punto

    except Exception as e:
        print(f"Ocurrió un error en def buscarPuntoUnico:{e}")
        return []


def procesar_actualizacion_form(data):
    try:
        with connectionBD() as conexion_PostgreSQLdb:
            with conexion_PostgreSQLdb.cursor(cursor_factory=DictCursor) as cursor:
                region = data.form['region']
                departamento = data.form['departamento']
                municipio = data.form['municipio']
                cod_hex = data.form['cod_hex']
                sup_parcelas = data.form['sup_parcelas']
                num_parcelas = data.form['num_parcelas']
                sup_muestra = data.form['sup_muestra']
                num_muestra = data.form['num_muestra']
                id = data.form['id']
                
                querySQL = """
                        UPDATE parcelas_registro
                        SET 
                            region = %s,
                            departamento = %s,
                            municipio = %s,
                            cod_hex = %s,
                            sup_parcelas = %s,
                            num_parcelas = %s,
                            sup_muestra = %s,
                            num_muestra = %s
                        WHERE id = %s
                """
                values = (region, departamento, municipio, cod_hex, sup_parcelas, num_parcelas, sup_muestra, num_muestra, id)

                cursor.execute(querySQL, values)
                conexion_PostgreSQLdb.commit()
        return cursor.rowcount or []
    except Exception as e:
        print(f"Ocurrió un error en procesar_actualizacion_form:{e}")
        return None

