# Python con PostgreSQL
# Para subir archivo tipo foto al servidor
from werkzeug.utils import secure_filename
import uuid  # Modulo de python para crear un string

from conexion.conexionBD import connectionBD  # Conexión a BD


import datetime
import re
import os

from os import remove  # Modulo  para remover archivo
from os import path  # Modulo para obtener la ruta o directorio

import openpyxl  # Para generar el excel
# biblioteca o modulo send_file para forzar la descarga
from flask import send_file

# Importar DictCursor para obtener resultados como diccionarios
# from psycopg2.extras import DictCursor
from psycopg2.extras import DictCursor  # Asegúrate de importar DictCursor

import random


# ********************************************************** sección de hexagonos **********************************************************

def procesar_imagen_perfil(foto):
    try:
        # Nombre original del archivo
        filename = secure_filename(foto.filename)
        extension = os.path.splitext(filename)[1]

        # Creando un string de 50 caracteres
        nuevoNameFile = (uuid.uuid4().hex + uuid.uuid4().hex)[:100]
        nombreFile = nuevoNameFile + extension

        # Construir la ruta completa de subida del archivo
        basepath = os.path.abspath(os.path.dirname(__file__))
        upload_dir = os.path.join(basepath, f'../static/fotos_hexagonos/')

        # Validar si existe la ruta y crearla si no existe
        if not os.path.exists(upload_dir):
            os.makedirs(upload_dir)
            # Dando permiso a la carpeta
            os.chmod(upload_dir, 0o755)

        # Construir la ruta completa de subida del archivo
        upload_path = os.path.join(upload_dir, nombreFile)
        foto.save(upload_path)

        return nombreFile

    except Exception as e:
        print("Error al procesar archivo:", e)
        return []


# Lista de hexagonos - PostgreSQL 
def sql_lista_hexagonosBD():
    try:
        with connectionBD() as conexion_PostgreSQLdb:
            with conexion_PostgreSQLdb.cursor(cursor_factory=DictCursor) as cursor:
                querySQL = """
                    SELECT 
                        up.id,
                        up.region, 
                        up.departamento,
                        up.municipio,
                        up.cod_hex,
                        up.sup_parcelas,
                        up.num_parcelas,
                        up.densidad,
                        up.sup_muestra,                      
                        up.num_muestra,
                        CASE
                            WHEN up.seleccionado = 1 THEN 'Seleccionado'
                            ELSE 'No Seleccionado'
                        END AS seleccion_hexagono
                    FROM unodc_parcelas AS up
                    WHERE up.seleccionado = 1
                    ORDER BY up.id DESC
                """
                cursor.execute(querySQL)
                hexagonosBD = cursor.fetchall()
        return hexagonosBD
    except Exception as e:
        print(f"Error en la función sql_lista_hexagonosBD:{e}")
        return None


# Detalles del hexagono - PostgreSQL 
def sql_detalles_hexagonosBD(idhexagono):
    try:
        with connectionBD() as conexion_PostgreSQLdb:
            with conexion_PostgreSQLdb.cursor(cursor_factory=DictCursor) as cursor:
                querySQL = """
                    SELECT
                        up.id,
                        up.region, 
                        up.departamento,
                        up.municipio,
                        up.cod_hex,
                        up.sup_parcelas,
                        up.num_parcelas,
                        up.densidad,
                        up.sup_muestra,                      
                        up.num_muestra,
                        CASE
                            WHEN up.seleccionado = 1 THEN 'Seleccionado'
                            ELSE 'No Seleccionado'
                        END AS seleccion_hexagono
                    FROM unodc_parcelas AS up 
                    WHERE id = %s
                    ORDER BY up.id DESC
                """
                cursor.execute(querySQL, (idhexagono,))
                hexagonosBD = cursor.fetchone()
        return hexagonosBD
    except Exception as e:
        print(f"Error en la función sql_detalles_hexagonosBD:{e}")
        return None


# Funcion hexagonos Informe (Reporte) - PostgreSQL
def hexagonosReporte():
    try:
        with connectionBD() as conexion_PostgreSQLdb:
            with conexion_PostgreSQLdb.cursor(cursor_factory=DictCursor) as cursor:
                querySQL = """
                    SELECT 
                        up.id,
                        up.region, 
                        up.departamento,
                        up.municipio,
                        up.cod_hex,
                        up.sup_parcelas,
                        up.num_parcelas,
                        up.densidad,
                        up.sup_muestra,                      
                        up.num_muestra,
                        CASE
                            WHEN up.seleccionado = 1 THEN 'Seleccionado'
                            ELSE 'No Seleccionado'
                        END AS seleccion_hexagono
                    FROM unodc_parcelas AS up 
                    WHERE up.seleccionado = 1
                    ORDER BY up.id DESC
                """
                cursor.execute(querySQL)
                hexagonosBD = cursor.fetchall()
        return hexagonosBD
    except Exception as e:
        print(f"Error en la función hexagonosReporte:{e}")
        return None

# Funcion hexagonos Generar Reporte Excel - PostgreSQL
def generarReporteExcelxhex():
    dataHexagonos = hexagonosReporte()
    wb = openpyxl.Workbook()
    hoja = wb.active

    # Agregar la fila de encabezado con los títulos
    cabeceraExcel = ( "Region", "Departamento", "Municipio", "Cod Hex", "Sup Parcelas", "Num Parcelas", "Densidad", "Sup Muestra", "Num Muestra", "Seleccionado")
                     

    hoja.append(cabeceraExcel)

    # Formato para números en moneda colombiana y sin decimales
    formato_moneda_colombiana = '#,##0'

    # Agregar los registros a la hoja
    for registro in dataHexagonos:
        region = registro['region']
        departamento = registro['departamento']
        municipio = registro['municipio']
        cod_hex = registro['cod_hex']
        sup_parcelas = registro['sup_parcelas']
        num_parcelas = registro['num_parcelas']
        densidad = registro['densidad']
        sup_muestra = registro['sup_muestra']
        num_muestra = registro['num_muestra']
        seleccion_hexagono = registro['seleccion_hexagono']
        
        # Agregar los valores a la hoja
        hoja.append((region, departamento, municipio, cod_hex, sup_parcelas, num_parcelas, densidad, sup_muestra, num_muestra, seleccion_hexagono))

        # Itera a través de las filas y aplica el formato a la columna G
        for fila_num in range(2, hoja.max_row + 1):
            columna = 7  # Columna G
            celda = hoja.cell(row=fila_num, column=columna)
            celda.number_format = formato_moneda_colombiana

    fecha_actual = datetime.datetime.now()
    archivoExcel = f"Reporte_hexagonos_{fecha_actual.strftime('%Y_%m_%d')}.xlsx"
    carpeta_descarga = "../static/downloads-excel"
    ruta_descarga = os.path.join(os.path.dirname(
        os.path.abspath(__file__)), carpeta_descarga)

    if not os.path.exists(ruta_descarga):
        os.makedirs(ruta_descarga)
        # Dando permisos a la carpeta
        os.chmod(ruta_descarga, 0o755)

    ruta_archivo = os.path.join(ruta_descarga, archivoExcel)
    wb.save(ruta_archivo)

    # Enviar el archivo como respuesta HTTP
    return send_file(ruta_archivo, as_attachment=True)

# Buscar hexagonos Unico - PostgreSQL 
def buscarHexagonoBD(search):
    try:
        with connectionBD() as conexion_PostgreSQLdb:
            with conexion_PostgreSQLdb.cursor(cursor_factory=DictCursor) as cursor:
                querySQL = """
                    SELECT 
                        up.id,
                        up.region, 
                        up.departamento,
                        up.municipio,
                        up.cod_hex,
                        up.sup_parcelas,
                        up.num_parcelas,
                        up.densidad,
                        up.sup_muestra,                      
                        up.num_muestra,
                        CASE
                            WHEN up.seleccionado = 1 THEN 'Seleccionado'
                            ELSE 'No Seleccionado'
                        END AS seleccion_hexagono
                    FROM unodc_parcelas AS up 
                    WHERE up.cod_hex ILIKE %s 
                    ORDER BY up.id DESC
                """
                search_pattern = f"%{search}%"  # Agregar "%" alrededor del término de búsqueda
                cursor.execute(querySQL, (search_pattern,))
                resultado_busqueda = cursor.fetchall()
                return resultado_busqueda

    except Exception as e:
        print(f"Ocurrió un error en def buscarHexagonoBD:{e}")
        return []

# Buscar hexagono Unico (cuando viene del grid) - PostgreSQL XXXXXXXXXXXXXX
def buscarHexagonoUnico(id):
    try:
        with connectionBD() as conexion_PostgreSQLdb:
            with conexion_PostgreSQLdb.cursor(cursor_factory=DictCursor) as cursor:
                querySQL = """
                    SELECT 
                        up.id,
                        up.region, 
                        up.departamento,
                        up.municipio,
                        up.cod_hex,
                        up.sup_parcelas,
                        up.num_parcelas,
                        up.densidad,
                        up.sup_muestra,                      
                        up.num_muestra,
                        CASE
                            WHEN up.seleccionado = 1 THEN 'Seleccionado'
                            ELSE 'No Seleccionado'
                        END AS seleccion_hexagono
                    FROM unodc_parcelas AS up 
                    WHERE up.id = %s LIMIT 1
                """
                cursor.execute(querySQL, (id,))
                hexagono = cursor.fetchone()
                return hexagono

    except Exception as e:
        print(f"Ocurrió un error en def buscarHexagonoUnico:{e}")
        return []


def procesar_actualizacion_form(data):
    try:
        with connectionBD() as conexion_PostgreSQLdb:
            with conexion_PostgreSQLdb.cursor(cursor_factory=DictCursor) as cursor:
                region = data.form['region']
                departamento = data.form['departamento']
                municipio = data.form['municipio']
                cod_hex = data.form['cod_hex']
                sup_parcelas = data.form['sup_parcelas']
                num_parcelas = data.form['num_parcelas']
                sup_muestra = data.form['sup_muestra']
                num_muestra = data.form['num_muestra']
                id = data.form['id']
                
                querySQL = """
                        UPDATE unodc_parcelas
                        SET 
                            region = %s,
                            departamento = %s,
                            municipio = %s,
                            cod_hex = %s,
                            sup_parcelas = %s,
                            num_parcelas = %s,
                            sup_muestra = %s,
                            num_muestra = %s
                        WHERE id = %s
                """
                values = (region, departamento, municipio, cod_hex, sup_parcelas, num_parcelas, sup_muestra, num_muestra, id)

                cursor.execute(querySQL, values)
                conexion_PostgreSQLdb.commit()
        return cursor.rowcount or []
    except Exception as e:
        print(f"Ocurrió un error en procesar_actualizacion_form:{e}")
        return None


# generar_muestras_aleatorias - PostgreSQL
def generar_muestras_aleatorias():
    try:
        with connectionBD() as conexion_PostgreSQLdb:
            with conexion_PostgreSQLdb.cursor(cursor_factory=DictCursor) as cursor:
                # Paso inicial: vaciar sorteos en 0
                cursor.execute("""
                    UPDATE unodc_parcelas 
                    SET sorteado = 0, seleccionado = 0, sup_muestra = 0, num_muestra = 0 
                    WHERE true
                """)
                conexion_PostgreSQLdb.commit()

                # Recorrer municipios seleccionados
                cursor.execute("SELECT * FROM unodc_municipios WHERE seleccionado = 1")
                municipios = cursor.fetchall()

                for municipio in municipios:
                    cod_mun = municipio['cod_mun']
                    sup_muestra_mun = municipio['sup_muestra']

                    # Inicializamos la variable de la suma de las parcelas seleccionadas
                    suma_par = 0

                    # Recorremos las parcelas del municipio ordenadas por superficie y densidad
                    cursor.execute("""
                        SELECT * FROM unodc_parcelas 
                        WHERE cod_mun = %s 
                        ORDER BY densidad DESC, sup_parcelas
                    """, (cod_mun,))
                    parcelas = cursor.fetchall()

                    # Creamos un array para almacenar las parcelas con sus respectivos pesos
                    parcelas_con_pesos = []

                    # Almacenamos las parcelas y su peso basado en la densidad
                    for parcela in parcelas:
                        sup_parcelas = parcela['sup_parcelas']
                        densidad = parcela['densidad']
                        id_parcela = parcela['id']

                        # Asignar peso según la densidad
                        if densidad == 3:       # Alta densidad
                            peso = 30
                        elif densidad == 2:     # Media densidad
                            peso = 15
                        elif densidad == 1:     # Baja Densidad 
                            peso = 1

                        # Añadir la parcela y su peso al array
                        parcelas_con_pesos.append({
                            'id': id_parcela,
                            'sup_parcelas': sup_parcelas,
                            'peso': peso
                        })

                    # Empezamos a seleccionar las parcelas
                    while len(parcelas_con_pesos) > 0:
                        # Elegimos una parcela aleatoriamente, ponderada por su peso
                        total_peso = sum(parcela['peso'] for parcela in parcelas_con_pesos)
                        random_peso = random.randint(1, total_peso)

                        suma_peso = 0
                        seleccionada = None

                        # Buscamos la parcela cuya ponderación incluye el número aleatorio generado
                        for parcela in parcelas_con_pesos:
                            suma_peso += parcela['peso']
                            if suma_peso >= random_peso:
                                seleccionada = parcela
                                break

                        # Verificamos si podemos seleccionar la parcela sin exceder la muestra
                        if suma_par + seleccionada['sup_parcelas'] <= sup_muestra_mun:
                            # Actualizamos la parcela seleccionada
                            cursor.execute("""
                                UPDATE unodc_parcelas 
                                SET sorteado = sorteado + 1, 
                                    sup_muestra = %s, 
                                    num_muestra = num_parcelas, 
                                    seleccionado = 1 
                                WHERE id = %s
                            """, (seleccionada['sup_parcelas'], seleccionada['id']))
                            conexion_PostgreSQLdb.commit()

                            # Acumulamos la superficie seleccionada
                            suma_par += seleccionada['sup_parcelas']

                        # Si ya hemos alcanzado el límite de la muestra, salimos del ciclo
                        if suma_par >= sup_muestra_mun:
                            break

                        # Eliminar la parcela seleccionada de las opciones disponibles
                        parcelas_con_pesos = [parcela for parcela in parcelas_con_pesos if parcela['id'] != seleccionada['id']]

                # Actualizamos el shape "region22xls-point" en POSTGIS
                cursor.execute("UPDATE parcelas SET seleccionado = 0 WHERE TRUE")
                cursor.execute("""
                    UPDATE parcelas 
                    SET seleccionado = 1
                    FROM unodc_parcelas
                    WHERE parcelas.cod_hex = unodc_parcelas.cod_hex AND unodc_parcelas.seleccionado = 1
                """)
                conexion_PostgreSQLdb.commit()

                # Contar cuántos registros fueron seleccionados
                cursor.execute("SELECT COUNT(*) AS total_seleccionados FROM unodc_parcelas WHERE seleccionado = 1")
                resultado = cursor.fetchone()
                total_seleccionados = resultado['total_seleccionados']

        return total_seleccionados  # Devuelve el número de registros seleccionados
    except Exception as e:
        print(f"Error en la función generar_muestras_aleatorias: {e}")
        return False  # Si ocurre un error
    